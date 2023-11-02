import os
import openpyxl

# Define the path to the txt files
txt_path = r"C:\Users\keiss\OneDrive\Desktop\she_codes\python_thesis"

# Get a list of all txt files in the directory
txt_files = [f for f in os.listdir(txt_path) if f.endswith('.txt')]

# Create an empty dictionary to hold the data
data_dict = {}

# Loop through the txt files and extract the data
for txt_file in txt_files:
    with open(os.path.join(txt_path, txt_file), 'r') as f:
        for line in f:
            if 'ID ' in line:
                id_code = line.split('ID ')[1][:8]
                date_time = line.split(',')[1].strip()
                # remove "on MRI-PCO" from date_time string
                date_time = date_time.split(' on ')[0]
                if id_code in data_dict:
                    data_dict[id_code].append(date_time)
                else:
                    data_dict[id_code] = [date_time]
                    
# Create a new Excel workbook
wb = openpyxl.Workbook()

# Get the active sheet
ws = wb.active

# Add headers to the worksheet
ws['A1'] = 'ID Code'
ws['B1'] = 'Date/Time (File 1)'
ws['C1'] = 'Date/Time (File 2)'
ws['D1'] = 'Date/Time (File 3)'

# Loop through the id codes and dates/times and add them to the worksheet
row_num = 2
for id_code, date_times in data_dict.items():
    ws.cell(row=row_num, column=1).value = id_code
    for i, dt in enumerate(date_times):
        ws.cell(row=row_num, column=i+2).value = dt
    row_num += 1

# Save the workbook
wb.save("C:\\Users\\keiss\\OneDrive\\Desktop\\she_codes\\python_thesis\\subject_runs.xlsx")
